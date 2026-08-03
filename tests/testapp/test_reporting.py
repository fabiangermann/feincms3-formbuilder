import io

from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import User
from django.test import RequestFactory, TestCase
from openpyxl import load_workbook

from feincms3_formbuilder.admin import BaseFormSubmissionAdmin, make_export_action
from feincms3_formbuilder.reporting import _sheet_title, build_submissions_xlsx
from testapp.admin import FormSubmissionAdmin
from testapp.models import (
    ConfiguredForm,
    FormStep,
    FormSubmission,
    Text,
)
from testapp.renderer import renderer


def _rows(response):
    """Read the single-sheet workbook back into a list of value rows."""
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook[workbook.sheetnames[0]]
    return workbook, [list(row) for row in sheet.iter_rows(values_only=True)]


# The four leading metadata columns every export starts with.
METADATA_HEADER = ["ID", "Submitted at", "IP address", "User agent"]


class ExportSubmissionsTest(TestCase):
    """End-to-end coverage of the admin export action and workbook contents."""

    def setUp(self):
        self.factory = RequestFactory()
        self.export = make_export_action(renderer)
        self.form = ConfiguredForm.objects.create(
            name="Contact", slug="contact", form_type="simple",
        )
        Text.objects.create(
            parent=self.form, region="form", ordering=10,
            name="name", label="Name", is_required=True,
        )

    def _export(self, queryset):
        request = self.factory.get("/")
        return self.export(None, request, queryset)

    def test_export_single_submission(self):
        FormSubmission.objects.create(
            configured_form=self.form,
            data={"name": "Alice"},
            ip_address="127.0.0.1",
            user_agent="TestAgent/1.0",
        )
        response = self._export(FormSubmission.objects.all())

        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="form-submissions.xlsx"',
        )
        workbook, rows = _rows(response)
        self.assertEqual(workbook.sheetnames, ["Contact"])
        self.assertEqual(rows[0], METADATA_HEADER + ["Name"])
        self.assertEqual(rows[1][2], "127.0.0.1")
        self.assertEqual(rows[1][3], "TestAgent/1.0")
        self.assertEqual(rows[1][4], "Alice")

    def test_export_multiple_submissions(self):
        FormSubmission.objects.create(configured_form=self.form, data={"name": "Alice"})
        FormSubmission.objects.create(configured_form=self.form, data={"name": "Bob"})
        _workbook, rows = _rows(self._export(FormSubmission.objects.all()))

        # One header row plus one row per submission, all on a single sheet.
        self.assertEqual(len(rows), 3)
        self.assertEqual({rows[1][4], rows[2][4]}, {"Alice", "Bob"})

    def test_export_multiple_forms(self):
        other = ConfiguredForm.objects.create(
            name="Newsletter", slug="newsletter", form_type="simple",
        )
        Text.objects.create(
            parent=other, region="form", ordering=10,
            name="email", label="Email", is_required=True,
        )
        FormSubmission.objects.create(configured_form=self.form, data={"name": "Alice"})
        FormSubmission.objects.create(configured_form=other, data={"email": "b@x.test"})

        workbook = load_workbook(
            io.BytesIO(self._export(FormSubmission.objects.all()).content)
        )
        self.assertEqual(sorted(workbook.sheetnames), ["Contact", "Newsletter"])


class ExportColumnOrderingTest(TestCase):
    """Field columns must follow region order, then ``ordering`` within a region.

    This is the regression guard for the multi-step export reading in the order
    the user filled the form in.
    """

    def test_columns_follow_form_order(self):
        form = ConfiguredForm.objects.create(
            name="Signup", slug="signup", form_type="multistep",
        )
        step1 = FormStep.objects.create(
            configured_form=form, title="Step 1", identifier="one", ordering=10,
        )
        step2 = FormStep.objects.create(
            configured_form=form, title="Step 2", identifier="two", ordering=20,
        )
        # ordering values deliberately interleave across the two regions.
        Text.objects.create(parent=form, region=step1.region_key, ordering=10, name="f_1c", label="Step 1 C")
        Text.objects.create(parent=form, region=step1.region_key, ordering=20, name="f_1a", label="Step 1 A")
        Text.objects.create(parent=form, region=step1.region_key, ordering=30, name="f_1b", label="Step 1 B")
        Text.objects.create(parent=form, region=step2.region_key, ordering=15, name="f_2b", label="Step 2 B")
        Text.objects.create(parent=form, region=step2.region_key, ordering=25, name="f_2a", label="Step 2 A")

        FormSubmission.objects.create(
            configured_form=form,
            data={
                "f_1c": "c1", "f_1a": "a1", "f_1b": "b1",
                "f_2b": "b2", "f_2a": "a2",
            },
        )

        xlsx = build_submissions_xlsx(FormSubmission.objects.all(), renderer=renderer)
        response = xlsx.to_response("form-submissions.xlsx")
        _workbook, rows = _rows(response)

        self.assertEqual(
            rows[0][len(METADATA_HEADER):],
            ["Step 1 C", "Step 1 A", "Step 1 B", "Step 2 B", "Step 2 A"],
        )
        self.assertEqual(
            rows[1][len(METADATA_HEADER):],
            ["c1", "a1", "b1", "b2", "a2"],
        )


class ExportActionWiringTest(TestCase):
    """The export action is opt-in: the base admin carries none, and adding the
    factory action to a subclass's ``actions`` is what surfaces it."""

    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_superuser("admin", "a@x.test", "pw")

    def _actions(self, admin_cls):
        request = self.factory.get("/")
        request.user = self.user
        return admin_cls(FormSubmission, AdminSite()).get_actions(request)

    def test_action_absent_from_base_admin(self):
        self.assertNotIn("export_submissions_xlsx", self._actions(BaseFormSubmissionAdmin))

    def test_action_present_when_added(self):
        self.assertIn("export_submissions_xlsx", self._actions(FormSubmissionAdmin))

    def test_factory_sets_action_label(self):
        self.assertEqual(
            str(make_export_action(renderer).short_description),
            "Export selected submissions to Excel",
        )


class SheetTitleTest(TestCase):
    """``_sheet_title`` keeps titles within openpyxl's constraints."""

    def test_strips_characters_excel_forbids(self):
        title = _sheet_title(r"a[b]c:d*e?f/g\h", set())
        self.assertFalse(set(title) & set(r"[]:*?/" + "\\"))

    def test_truncates_to_31_characters(self):
        self.assertEqual(len(_sheet_title("x" * 40, set())), 31)

    def test_disambiguates_duplicate_titles(self):
        used = set()
        self.assertEqual(_sheet_title("Contact", used), "Contact")
        self.assertEqual(_sheet_title("Contact", used), "Contact (2)")
